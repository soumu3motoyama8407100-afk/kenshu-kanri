from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "職員名簿"

GOLD="C89A55"; LIGHT_GOLD="FDF6EC"; LIGHT_GRAY="F3F4F6"; DARK="4A3020"; WHITE="FFFFFF"

def bd():
    s=Side(style="thin",color="E8D5B0")
    return Border(left=s,right=s,top=s,bottom=s)

# Row1: タイトル
ws.merge_cells("A1:M1")
ws["A1"]="職員名簿テンプレート（研修管理システム インポート用）"
ws["A1"].font=Font(name="メイリオ",bold=True,size=13,color=WHITE)
ws["A1"].fill=PatternFill("solid",fgColor=GOLD)
ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
ws.row_dimensions[1].height=28

# Row2: 注意書き
ws.merge_cells("A2:M2")
ws["A2"]="※ ヘッダー行（3行目）は変更しないでください  ／  ★の列は必須  ／  インポート前に「CSV UTF-8（コンマ区切り）」形式で保存してください"
ws["A2"].font=Font(name="メイリオ",size=9,color="DC2626")
ws["A2"].fill=PatternFill("solid",fgColor="FEF2F2")
ws["A2"].alignment=Alignment(horizontal="left",vertical="center")
ws.row_dimensions[2].height=18

# Row3: ヘッダー（13列）
headers=[
    "職員ID ★","パスワード ★","姓（苗字）★","名前 ★",
    "入社日","役職名","職員区分","所属",
    "管理部署","保有資格","認定研修","部署長","在籍状態"
]
for i,h in enumerate(headers,1):
    c=ws.cell(row=3,column=i,value=h)
    c.font=Font(name="メイリオ",bold=True,size=10,color=WHITE)
    c.fill=PatternFill("solid",fgColor=GOLD)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=bd()
ws.row_dimensions[3].height=22

# Row4: 入力説明
tips=[
    "例: 001, E001, 158",
    "例: pass001",
    "例: 山田",
    "例: 太郎",
    "例: 2020-04-01",
    "例: 主任, 管理者",
    "例: 正職員, パート\n非常勤, 派遣",
    "例: ホーム, デイ",
    "複数は|区切り\n例: ホーム|デイ",
    "複数は|区切り\n例: 介護福祉士|社会福祉士",
    "複数は|区切り\n例: 実務者研修",
    "部署長は「1」\nそれ以外は空欄",
    "退職者は「退職」\nそれ以外は空欄",
]
for i,t in enumerate(tips,1):
    c=ws.cell(row=4,column=i,value=t)
    c.font=Font(name="メイリオ",size=8,color="6B7280",italic=True)
    c.fill=PatternFill("solid",fgColor=LIGHT_GRAY)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    c.border=bd()
ws.row_dimensions[4].height=40

# Row5〜: サンプルデータ
samples=[
    ["E001","pass001","山田","太郎","2018-04-01","主任","正職員","ホーム","ホーム","介護福祉士","実務者研修","1",""],
    ["E002","pass002","鈴木","花子","2020-06-01","","正職員","ホーム","","介護職員初任者研修","","",""],
    ["E003","pass003","田中","次郎","2019-04-01","管理者","正職員","デイ","デイ","介護福祉士|社会福祉士","認知症介護実践者研修","1",""],
    ["E004","pass004","佐藤","美咲","2022-04-01","","パート","デイ","","介護職員初任者研修","","",""],
    ["E005","pass005","高橋","健一","2021-09-01","","正職員","小規模","","介護福祉士","","",""],
]
for r,row in enumerate(samples,5):
    fill=LIGHT_GOLD if r%2==1 else WHITE
    for col,val in enumerate(row,1):
        c=ws.cell(row=r,column=col,value=val)
        c.font=Font(name="メイリオ",size=10,color=DARK)
        c.fill=PatternFill("solid",fgColor=fill)
        c.alignment=Alignment(horizontal="left",vertical="center")
        c.border=bd()
    ws.row_dimensions[r].height=20

# 列幅
widths=[10,12,12,12,13,12,12,12,14,22,20,8,10]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width=w

out=r"C:\Users\heart42\Desktop\職員名簿テンプレート.xlsx"
wb.save(out)
print("保存完了:", out)
