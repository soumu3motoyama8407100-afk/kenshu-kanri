from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "職員名簿"

GOLD = "C89A55"
LIGHT_GOLD = "FDF6EC"
LIGHT_GRAY = "F3F4F6"
DARK = "4A3020"
WHITE = "FFFFFF"

def bd():
    s = Side(style="thin", color="E8D5B0")
    return Border(left=s, right=s, top=s, bottom=s)

# Row1: タイトル
ws.merge_cells("A1:K1")
ws["A1"] = "職員名簿テンプレート（研修管理システム インポート用）"
ws["A1"].font = Font(name="メイリオ", bold=True, size=13, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=GOLD)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Row2: 注意書き
ws.merge_cells("A2:K2")
ws["A2"] = "※ ヘッダー行（3行目）は変更しないでください  ／  ★の列は必須  ／  インポート前にCSV（UTF-8）形式で保存してください"
ws["A2"].font = Font(name="メイリオ", size=9, color="DC2626")
ws["A2"].fill = PatternFill("solid", fgColor="FEF2F2")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

# Row3: ヘッダー
headers = ["職員ID ★","パスワード ★","氏名 ★","部署 ★","入社年月日","保有資格","認定研修","部署長","管理部署","在籍状態","役職名"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = Font(name="メイリオ", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bd()
ws.row_dimensions[3].height = 22

# Row4: 入力説明
tips = [
    "例: 001, E001, 158",
    "例: pass001",
    "例: 山田 太郎",
    "例: ホーム",
    "例: 2020-04-01",
    "複数は|区切り\n例: 介護福祉士|社会福祉士",
    "複数は|区切り\n例: 実務者研修",
    "部署長は「1」\nそれ以外は空欄",
    "複数は|区切り\n例: ホーム|デイ",
    "退職者は「退職」\nそれ以外は空欄",
    "例: 主任, 管理者",
]
for i, t in enumerate(tips, 1):
    c = ws.cell(row=4, column=i, value=t)
    c.font = Font(name="メイリオ", size=8, color="6B7280", italic=True)
    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = bd()
ws.row_dimensions[4].height = 36

# Row5〜: サンプルデータ
samples = [
    ["E001","pass001","山田 太郎","ホーム","2018-04-01","介護福祉士","実務者研修","1","ホーム","","主任"],
    ["E002","pass002","鈴木 花子","ホーム","2020-06-01","介護職員初任者研修","","","","",""],
    ["E003","pass003","田中 次郎","デイ","2019-04-01","介護福祉士|社会福祉士","認知症介護実践者研修","1","デイ","","管理者"],
    ["E004","pass004","佐藤 美咲","デイ","2022-04-01","介護職員初任者研修","","","","",""],
    ["E005","pass005","高橋 健一","小規模","2021-09-01","介護福祉士","","","","",""],
]
for r, row in enumerate(samples, 5):
    fill = LIGHT_GOLD if r % 2 == 1 else WHITE
    for col, val in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name="メイリオ", size=10, color=DARK)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bd()
    ws.row_dimensions[r].height = 20

# 列幅
for i, w in enumerate([10,12,14,12,14,22,22,8,14,10,12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet2: 記入ガイド ─────────────────────────────────────
ws2 = wb.create_sheet("記入ガイド")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 55

ws2.merge_cells("A1:B1")
ws2["A1"] = "記入ガイド"
ws2["A1"].font = Font(name="メイリオ", bold=True, size=12, color=WHITE)
ws2["A1"].fill = PatternFill("solid", fgColor=GOLD)
ws2["A1"].alignment = Alignment(horizontal="center")
ws2.row_dimensions[1].height = 24

guides = [
    ("列名", "説明"),
    ("職員ID ★", "重複しない番号。数字のみ・英数字どちらもOK（例: 001, E001, 158）"),
    ("パスワード ★", "本人のログインパスワード（例: pass001）"),
    ("氏名 ★", "フルネーム（例: 山田 太郎）"),
    ("部署 ★", "所属部署。表記を統一してください（例: ホーム）"),
    ("入社年月日", "YYYY-MM-DD形式（例: 2020-04-01）"),
    ("保有資格", "複数は | で区切る（例: 介護福祉士|社会福祉士）"),
    ("認定研修", "複数は | で区切る（例: 実務者研修|認知症ケア）"),
    ("部署長", "部署長なら「1」、それ以外は空欄のまま"),
    ("管理部署", "部署長が担当する部署。複数は | で区切る"),
    ("在籍状態", "退職者のみ「退職」と入力。在籍中は空欄"),
    ("役職名", "主任、管理者、副主任など（任意）"),
    ("", ""),
    ("⚠️ 重要注意事項", ""),
    ("保存形式", "「CSV UTF-8（コンマ区切り）」で保存してからインポート"),
    ("3行目ヘッダー", "ヘッダー行は削除しないこと（アプリが1行目をスキップして読み込む）"),
    ("部署名の表記", "「ホーム」と「ホーム棟」など表記ゆれに注意"),
    ("パスワード", "後から変更できないため、本人が覚えやすいものに設定"),
]
for r, (col, desc) in enumerate(guides, 2):
    c1 = ws2.cell(row=r, column=1, value=col)
    c2 = ws2.cell(row=r, column=2, value=desc)
    c1.font = Font(name="メイリオ", bold=(r == 2), size=9, color=DARK)
    c2.font = Font(name="メイリオ", size=9, color=DARK)
    if r == 2:
        c1.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        c2.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    if col.startswith("⚠️"):
        c1.font = Font(name="メイリオ", bold=True, size=9, color="DC2626")

out = r"C:\Users\heart42\Desktop\職員名簿テンプレート.xlsx"
wb.save(out)
print("保存完了:", out)
